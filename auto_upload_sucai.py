import time
import requests
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class UploadMaterial(object):
    def __init__(self):
        ctoken = 'UaU69wef-Z9Qh5O0'
        self.upload_url = "https://mesa2-pre.alipay.com/api/source/uploadByStream"
        self.save_url = f"https://mesa2-pre.alipay.com/mesa2api/needle/batchSaveMaterial?ctoken={ctoken}"
        self.get_url = f"https://mesa2-pre.alipay.com/mesa2api/needle/getMaterialPageList?ctoken={ctoken}"
        self.edit_url = f"https://mesa2-pre.alipay.com/mesa2api/needle/editMaterial?ctoken={ctoken}"
        self.create_group_url = f"https://mesa2-pre.alipay.com/mesa2api/needle/saveMaterialGroup?ctoken={ctoken}"
        self.get_group_id_url = f"https://mesa2-pre.alipay.com/mesa2api/needle/getMaterialGroupPageList?ctoken={ctoken}"
        self.cookie = self.get_cookie()
        self.upload_headers = {
            'accept': '*/*',
            'accept-language': 'zh-CN,zh;q=0.9',
            'origin': 'https://mesa2-pre.alipay.com',
            'priority': 'u=1, i',
            'referer': 'https://mesa2-pre.alipay.com/fmsmng/material?groupId=004885694100037211&groupName=%E7%90%86%E8%B4%A2%E6%88%98%E6%96%97%E5%8A%9B',
            'sec-ch-ua': '"Google Chrome";v="135", "Not-A.Brand";v="8", "Chromium";v="135"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-origin',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36',
            'Cookie': self.cookie
        }
        self.headers = {
            'accept': 'application/json',
            'accept-language': 'zh-CN,zh;q=0.9',
            'content-type': 'application/json',
            'origin': 'https://mesa2-pre.alipay.com',
            'priority': 'u=1, i',
            'referer': 'https://mesa2-pre.alipay.com/fmsmng/material?groupId=004885694100037211&groupName=%E7%90%86%E8%B4%A2%E6%88%98%E6%96%97%E5%8A%9B',
            'sec-ch-ua': '"Google Chrome";v="135", "Not-A.Brand";v="8", "Chromium";v="135"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-origin',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36',
            'x-csrf-token': 'UaU69wef-Z9Qh5O0',
            'Cookie': self.cookie
        }

    @staticmethod
    def get_cookie():
        with open('cookie.txt', 'r') as f:
            cookie = f.read()
        return cookie

    def get_group_id(self, group_name=''):
        get_payload = json.dumps({
            "postData": [
                {
                    "pageNo": 0,
                    "pageSize": 19,
                }
            ]
        })
        response = requests.request("POST", self.get_group_id_url, headers=self.headers, data=get_payload)
        if response.status_code == 200:
            if group_name:
                if response.json()['data']['pageResult'][0]['materialGroupName'] == group_name:
                    group_id = response.json()['data']['pageResult'][0]['materialGroupId']
                    return group_id
                else:
                    print('素材组创建失败啦')
                    return False
            return True
        else:
            print('COOKIE失效啦')
            return False

    def create_group(self, group_name='test'):
        # 创建素材组
        create_payload = json.dumps({
            "postData": {
                "materialGroupName": group_name,
                "creator": "游弘扬"
            }
        })
        response = requests.request("POST", self.create_group_url, headers=self.headers, data=create_payload)
        if response.status_code == 200:
            return True
        else:
            print('创建素材组失败啦')
            return False

    def upload(self, group_id, file_name='upload.png', file_path=r'D:\code\ant\su_cai_zu\images\img.png'):
        # 上传素材
        payload = {}
        files = [
            ('file', (file_name, open(file_path, 'rb'), 'image/png'))
        ]
        upload_response = requests.request("POST", self.upload_url, headers=self.upload_headers, data=payload,
                                           files=files)
        file_url, file_size = upload_response.json()['results'][0]['url'], upload_response.json()['results'][0]['size']

        # 录入素材组
        save_payload = json.dumps({
            "postData": [
                {
                    "groupId": group_id,
                    "materialName": file_name[:-4],
                    "fileSize": file_size,
                    "fileUrl": file_url,
                    "fileType": "png",
                    "fileDuration": 0,
                    "materialCover": file_url,
                    "creator": "游弘扬",
                    "emotionClassifyList": [
                        "PROCESS_EARN_0.8_TO_1.5"
                    ],
                    "materialStatus": "1"
                }
            ]
        })
        requests.request("POST", self.save_url, headers=self.headers, data=save_payload)

    def get_material_id(self, group_id):
        # 获取素材id
        get_payload = json.dumps({
            "postData": {
                "pageNo": 0,
                "pageSize": 20,
                "groupId": group_id
            }
        })
        get_response = requests.request("POST", self.get_url, headers=self.headers, data=get_payload)
        material_id = get_response.json()['data']['pageResult'][0]['materialId']
        return material_id

    def edit_material(self, material_id, material_name='test', tag='PROCESS_EARN_0.8_TO_1.5', title='title',
                      material_writing='materialWriting',
                      material_content='materialContent'):
        # 编辑素材
        payload = json.dumps({
            "postData": {
                "materialId": material_id,
                "materialName": material_name,
                "emotionClassifyList": [
                    tag
                ],
                "materialStatus": "1",
                "editor": "游弘扬",
                "title": title,
                "materialWriting": material_writing,
                "materialContent": material_content
            }
        })
        requests.request("POST", self.edit_url, headers=self.headers, data=payload)


class ReadFile(object):
    def __init__(self, file_path='sucai.xlsx'):
        # 加载Excel文件
        self.wb = load_workbook(filename=file_path)
        self.ws = self.wb.active
        self.max_row = self.ws.max_row
        self.max_col = self.ws.max_column

    def get_img(self, img_row, img_col=2):
        # 遍历工作表中的所有图片
        for idx, image in enumerate(self.ws._images):
            # 获取锚点的列和行索引（注意：列和行是从0开始计数）
            col = image.anchor._from.col
            row = image.anchor._from.row
            if row + 1 == img_row and col + 1 == img_col:
                image_data = image._data()
                image_path = f'images/img.png'
                with open(image_path, 'wb') as img_file:
                    img_file.write(image_data)

    def get_info(self, row_number, col_number):
        return self.ws.cell(row=row_number, column=col_number).value

    def get_row_info(self, row, img=True):
        name = self.get_info(row, 1)
        title = self.get_info(row, 3)
        material_writing = self.get_info(row, 4)
        material_content = self.get_info(row, 5)
        if img:
            self.get_img(row, 2)
        return name, title, material_writing, material_content


if __name__ == '__main__':
    RF = ReadFile('sucai.xlsx')
    group_name = '理财战斗力'
    has_img = True

    UM = UploadMaterial()
    tag_dict = {'赚0%-0.8%': 'PROCESS_EARN_0_TO_0.8', '赚0.8%-1.5%': 'PROCESS_EARN_0.8_TO_1.5',
                '赚1.5%-2.5%': 'PROCESS_EARN_1.5_TO_2.5', '赚2.5%-3.8%': 'PROCESS_EARN_2.5_TO_3.8',
                '赚3.8%-5.0%': 'PROCESS_EARN_3.8_TO_5.0', '赚5.0%-7.0%': 'PROCESS_EARN_5.0_TO_7.0',
                '赚7.0%-9.5%': 'PROCESS_EARN_7.0_TO_9.5', '赚9.5%-12%': 'PROCESS_EARN_9.5_TO_12',
                '赚12%-15%': 'PROCESS_EARN_12_TO_15', '赚15%-19%': 'PROCESS_EARN_15_TO_19',
                '赚19%-24%': 'PROCESS_EARN_19_TO_24', '赚24%-30%': 'PROCESS_EARN_24_TO_30',
                '赚30%-40%': 'PROCESS_EARN_30_TO_40', '赚40%-60%': 'PROCESS_EARN_40_TO_60',
                '赚60%以上': 'PROCESS_EARN_OVER_60', '亏0%至-2%': 'PROCESS_LOSE_0_TO_NEG_2',
                '亏-2%至-4%': 'PROCESS_LOSE_NEG_2_TO_NEG_4', '亏-4%至-6%': 'PROCESS_LOSE_NEG_4_TO_NEG_6',
                '亏-6%至-8%': 'PROCESS_LOSE_NEG_6_TO_NEG_8', '亏-8%至-10%': 'PROCESS_LOSE_NEG_8_TO_NEG_10',
                '亏-10%至-12%': 'PROCESS_LOSE_NEG_10_TO_NEG_12', '亏-12%至-15%': 'PROCESS_LOSE_NEG_12_TO_NEG_15',
                '亏-15%至-18%': 'PROCESS_LOSE_NEG_15_TO_NEG_18', '亏-18%至-21%': 'PROCESS_LOSE_NEG_18_TO_NEG_21',
                '亏-21%至-25%': 'PROCESS_LOSE_NEG_21_TO_NEG_25', '亏-25%至-30%': 'PROCESS_LOSE_NEG_25_TO_NEG_30',
                '亏-30%至-35%': 'PROCESS_LOSE_NEG_30_TO_NEG_35', '亏-35%至-45%': 'PROCESS_LOSE_NEG_35_TO_NEG_45',
                '亏-45%至-60%': 'PROCESS_LOSE_NEG_45_TO_NEG_60', '亏-60%以上': 'PROCESS_LOSE_OVER_NEG_60'}

    UM.create_group(group_name)
    group_id = UM.get_group_id(group_name)
    if UM.get_group_id():
        for row in range(2, RF.max_row + 1):
            print(f'开始上传第{row}行素材')
            name, title, material_writing, material_content = RF.get_row_info(row, img=has_img)
            tag = tag_dict[name]
            UM.upload(group_id)
            material_id = UM.get_material_id(group_id)
            UM.edit_material(material_id,
                             material_name=name,
                             tag=tag,
                             title=title,
                             material_writing=material_writing,
                             material_content=material_content)
            time.sleep(1)
